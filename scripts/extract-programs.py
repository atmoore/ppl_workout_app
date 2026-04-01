#!/usr/bin/env python3
"""
Extract workout program data from Jeff Nippard Excel spreadsheets.
Outputs standardized JSON files to scripts/seed-data/.
"""

import json
import re
import os
import datetime
from pathlib import Path

import openpyxl

JEFF_NIPPARD_DIR = Path("/Users/austinmoore/projects/Jeff Nippard")
OUTPUT_DIR = Path(__file__).parent / "seed-data"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def datetime_to_string(val):
    """Convert Excel-misinterpreted date back to 'M-D' string."""
    if isinstance(val, datetime.datetime):
        return f"{val.month}-{val.day}"
    return val


def clean_value(val):
    """Normalize a cell value: convert datetimes, strip strings."""
    val = datetime_to_string(val)
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
    return val


def clean_substitution(val):
    """Clean substitution name: replace newlines with spaces."""
    val = clean_value(val)
    if isinstance(val, str):
        val = val.replace("\n", " ").strip()
        if not val or val.upper() == "N/A":
            return None
    return val


def detect_workout_type(name: str) -> str:
    """Detect workout type from workout name."""
    name_upper = name.upper()
    if "PUSH" in name_upper:
        return "push"
    if "PULL" in name_upper:
        return "pull"
    if "LEG" in name_upper:
        return "legs"
    if "UPPER" in name_upper:
        return "upper"
    if "LOWER" in name_upper:
        return "lower"
    # Powerbuilding workouts
    if "FULL BODY" in name_upper or "FULL_BODY" in name_upper:
        return "full"
    # Any other powerbuilding label (CHEST, BACK, etc.)
    return "full"


def is_rest_day(val) -> bool:
    """Check if a row is a rest day marker."""
    if val is None:
        return False
    return "rest day" in str(val).lower()


def is_week_marker(val) -> bool:
    """Check if a cell value is a week marker like 'Week 1' or 'Week 1 Day 2'."""
    if val is None:
        return False
    return bool(re.match(r"^Week\s+\d+", str(val).strip(), re.IGNORECASE))


def is_header_row(row) -> bool:
    """Check if a row is a header row."""
    for cell in row:
        if cell and str(cell).strip().lower() in ("exercise", "exercises"):
            return True
    return False


def extract_week_number(val) -> int:
    """Extract week number from 'Week N' string."""
    m = re.search(r"(\d+)", str(val))
    return int(m.group(1)) if m else 1


def is_workout_label(val) -> bool:
    """Check if a cell is a workout label (Push #1, Pull #1, FULL BODY 1:, SQUAT TEST:, etc.)."""
    if val is None:
        return False
    s = str(val).strip()
    patterns = [
        r"^(Push|Pull|Legs|Leg|Upper|Lower)\s+#?\d+",
        r"^FULL\s+BODY\s+\d+",
        r"^(LOWER|UPPER|PUSH|PULL)\s+#?\d+",
        r"^FULL\s+BODY\s+\d+:",
        # Powerbuilding 3.0 test week labels
        r"^(SQUAT|BENCH|DEADLIFT)\s+TEST:",
    ]
    for p in patterns:
        if re.match(p, s, re.IGNORECASE):
            return True
    return False


def normalize_workout_label(val) -> str:
    """Strip trailing colon from workout labels."""
    return str(val).strip().rstrip(":")


# ---------------------------------------------------------------------------
# PPL parser (Ultimate PPL files)
# ---------------------------------------------------------------------------
# Column layout (0-indexed within values_only row):
#   A=0  B=1  C=2  D=3  E=4  F=5  G=6  H=7  I=8  J=9  K=10
#   workout_label, exercise, warmup_sets, working_sets, reps, load(skip),
#   rpe, rest, sub1, sub2, notes

def parse_ppl_sheet(ws, sheet_name: str) -> dict:
    """Parse a single PPL sheet into phase structure."""
    rows = list(ws.iter_rows(min_row=7, values_only=True))

    weeks = []
    current_week_number = None
    current_week_workouts = []
    current_workout = None
    exercise_order = 0

    for row in rows:
        col_a = clean_value(row[0]) if len(row) > 0 else None
        col_b = clean_value(row[1]) if len(row) > 1 else None

        # Skip header rows (but first check if col_a has a week marker on the same row)
        if is_header_row(row):
            if col_a and is_week_marker(col_a):
                week_num = extract_week_number(col_a)
                if current_week_number is None:
                    # Start first week
                    current_week_number = week_num
                elif week_num != current_week_number:
                    # New week — save current and start fresh
                    if current_workout is not None:
                        current_week_workouts.append(current_workout)
                        current_workout = None
                    weeks.append({
                        "week_number": current_week_number,
                        "workouts": current_week_workouts,
                    })
                    current_week_number = week_num
                    current_week_workouts = []
                    exercise_order = 0
                # Same week number (e.g. "Week 1 Day 2") — just continue within current week
            continue

        # Skip rest day rows
        if is_rest_day(col_a):
            continue

        # Week marker (standalone row, not a header row)
        if col_a and is_week_marker(col_a):
            week_num = extract_week_number(col_a)
            if current_week_number is None:
                current_week_number = week_num
            elif week_num != current_week_number:
                if current_workout is not None:
                    current_week_workouts.append(current_workout)
                    current_workout = None
                weeks.append({
                    "week_number": current_week_number,
                    "workouts": current_week_workouts,
                })
                current_week_number = week_num
                current_week_workouts = []
                exercise_order = 0
            # Same week number — continue (e.g. "Week 1 Day 2" repeat)
            continue

        # Workout label (col A has label, col B has first exercise)
        if col_a and is_workout_label(col_a):
            if current_workout is not None:
                current_week_workouts.append(current_workout)
            workout_name = normalize_workout_label(col_a)
            current_workout = {
                "name": workout_name,
                "type": detect_workout_type(workout_name),
                "day_number": len(current_week_workouts) + 1,
                "exercises": [],
            }
            exercise_order = 0

            # col_b may have the first exercise on this same row
            if col_b:
                exercise_order += 1
                exercise = build_ppl_exercise(row, exercise_order)
                if exercise:
                    current_workout["exercises"].append(exercise)
            continue

        # Regular exercise row (col_a is None, col_b has exercise name)
        if col_b and current_workout is not None:
            exercise_order += 1
            exercise = build_ppl_exercise(row, exercise_order)
            if exercise:
                current_workout["exercises"].append(exercise)

    # Flush remaining
    if current_workout is not None:
        current_week_workouts.append(current_workout)
    if current_week_number is not None:
        weeks.append({
            "week_number": current_week_number,
            "workouts": current_week_workouts,
        })

    return {
        "phase_number": None,  # Will be set by caller
        "name": sheet_name,
        "description": None,
        "weeks": weeks,
    }


def build_ppl_exercise(row, order: int) -> dict | None:
    """Build an exercise dict from a PPL row."""
    def get(idx):
        return clean_value(row[idx]) if len(row) > idx else None

    name = get(1)
    if not name:
        return None

    warmup = get(2)
    working = get(3)
    reps = get(4)
    # skip col 5 (load)
    rpe = get(6)
    rest = get(7)
    sub1 = clean_substitution(row[8]) if len(row) > 8 else None
    sub2 = clean_substitution(row[9]) if len(row) > 9 else None
    notes = get(10)

    substitutions = []
    if sub1 and sub1.upper() != "N/A":
        substitutions.append({"option_number": 1, "name": sub1})
    if sub2 and sub2.upper() != "N/A":
        substitutions.append({"option_number": 2, "name": sub2})

    # working_sets should be int if possible
    if working is not None:
        try:
            working = int(working)
        except (ValueError, TypeError):
            pass

    return {
        "order": order,
        "name": str(name),
        "warmup_sets": str(warmup) if warmup is not None else None,
        "working_sets": working,
        "reps": str(reps) if reps is not None else None,
        "rpe": str(rpe) if rpe is not None else None,
        "rest": str(rest) if rest is not None else None,
        "notes": str(notes) if notes is not None else None,
        "substitutions": substitutions,
    }


# ---------------------------------------------------------------------------
# Powerbuilding 2.0 parser
# ---------------------------------------------------------------------------
# Column layout (0-indexed):
#   A=0  B=1  C=2  D=3  E=4  F=5  G=6  H=7  I=8  J=9  K=10
#   (unused), week/workout_label, exercise, warmup_sets, working_sets,
#   reps, load(skip), %1RM(skip), rpe, rest, notes

def parse_pb2_sheet(ws, sheet_name: str) -> dict:
    """Parse a Powerbuilding 2.0 sheet."""
    rows = list(ws.iter_rows(min_row=13, values_only=True))

    weeks = []
    current_week_number = None
    current_week_workouts = []
    current_workout = None
    exercise_order = 0

    for row in rows:
        col_b = clean_value(row[1]) if len(row) > 1 else None
        col_c = clean_value(row[2]) if len(row) > 2 else None

        # Skip header rows
        if is_header_row(row):
            if col_b and is_week_marker(col_b):
                week_num = extract_week_number(col_b)
                if current_week_number is None:
                    current_week_number = week_num
                elif week_num != current_week_number:
                    if current_workout is not None:
                        current_week_workouts.append(current_workout)
                        current_workout = None
                    weeks.append({
                        "week_number": current_week_number,
                        "workouts": current_week_workouts,
                    })
                    current_week_number = week_num
                    current_week_workouts = []
                    exercise_order = 0
            continue

        # Skip rest days
        if is_rest_day(col_b):
            continue

        # Week marker in col B
        if col_b and is_week_marker(col_b):
            week_num = extract_week_number(col_b)
            if current_week_number is None:
                current_week_number = week_num
            elif week_num != current_week_number:
                if current_workout is not None:
                    current_week_workouts.append(current_workout)
                    current_workout = None
                weeks.append({
                    "week_number": current_week_number,
                    "workouts": current_week_workouts,
                })
                current_week_number = week_num
                current_week_workouts = []
                exercise_order = 0
            continue

        # Workout label in col B
        if col_b and is_workout_label(col_b):
            if current_workout is not None:
                current_week_workouts.append(current_workout)
            workout_name = normalize_workout_label(col_b)
            current_workout = {
                "name": workout_name,
                "type": detect_workout_type(workout_name),
                "day_number": len(current_week_workouts) + 1,
                "exercises": [],
            }
            exercise_order = 0

            # col_c may have first exercise
            if col_c:
                exercise_order += 1
                exercise = build_pb2_exercise(row, exercise_order)
                if exercise:
                    current_workout["exercises"].append(exercise)
            continue

        # Regular exercise row
        if col_c and current_workout is not None:
            exercise_order += 1
            exercise = build_pb2_exercise(row, exercise_order)
            if exercise:
                current_workout["exercises"].append(exercise)

    # Flush
    if current_workout is not None:
        current_week_workouts.append(current_workout)
    if current_week_number is not None:
        weeks.append({
            "week_number": current_week_number,
            "workouts": current_week_workouts,
        })

    return {
        "phase_number": 1,
        "name": sheet_name,
        "description": None,
        "weeks": weeks,
    }


def build_pb2_exercise(row, order: int) -> dict | None:
    """Build exercise from PB 2.0 row (col offset by 1)."""
    def get(idx):
        return clean_value(row[idx]) if len(row) > idx else None

    name = get(2)
    if not name:
        return None

    warmup = get(3)
    working = get(4)
    reps = get(5)
    # col 6 = load (skip), col 7 = %1RM (skip)
    rpe = get(8)
    rest = get(9)
    notes = get(10)

    if working is not None:
        try:
            working = int(working)
        except (ValueError, TypeError):
            pass

    return {
        "order": order,
        "name": str(name),
        "warmup_sets": str(warmup) if warmup is not None else None,
        "working_sets": working,
        "reps": str(reps) if reps is not None else None,
        "rpe": str(rpe) if rpe is not None else None,
        "rest": str(rest) if rest is not None else None,
        "notes": str(notes) if notes is not None else None,
        "substitutions": [],
    }


# ---------------------------------------------------------------------------
# Powerbuilding 3.0 parser
# ---------------------------------------------------------------------------
# Column layout (0-indexed):
#   A=0  B=1  C=2  D=3  E=4  F=5  G=6  H=7  I=8  J=9
#   workout_label, exercise, warmup_sets, working_sets, reps, load(skip),
#   %1RM(skip), rpe, rest, notes

def parse_pb3_sheet(ws, sheet_name: str) -> dict:
    """Parse PB 3.0 sheet (header at row 11, data from row 12)."""
    rows = list(ws.iter_rows(min_row=11, values_only=True))

    weeks = []
    current_week_number = None
    current_week_workouts = []
    current_workout = None
    exercise_order = 0

    for row in rows:
        col_a = clean_value(row[0]) if len(row) > 0 else None
        col_b = clean_value(row[1]) if len(row) > 1 else None

        # Skip header rows
        if is_header_row(row):
            if col_a and is_week_marker(col_a):
                week_num = extract_week_number(col_a)
                if current_week_number is None:
                    current_week_number = week_num
                elif week_num != current_week_number:
                    if current_workout is not None:
                        current_week_workouts.append(current_workout)
                        current_workout = None
                    weeks.append({
                        "week_number": current_week_number,
                        "workouts": current_week_workouts,
                    })
                    current_week_number = week_num
                    current_week_workouts = []
                    exercise_order = 0
            continue

        # Skip rest days
        if is_rest_day(col_a) or is_rest_day(col_b):
            continue

        # Week marker
        if col_a and is_week_marker(col_a):
            week_num = extract_week_number(col_a)
            if current_week_number is None:
                current_week_number = week_num
            elif week_num != current_week_number:
                if current_workout is not None:
                    current_week_workouts.append(current_workout)
                    current_workout = None
                weeks.append({
                    "week_number": current_week_number,
                    "workouts": current_week_workouts,
                })
                current_week_number = week_num
                current_week_workouts = []
                exercise_order = 0
            continue

        # Workout label in col A
        if col_a and is_workout_label(col_a):
            if current_workout is not None:
                current_week_workouts.append(current_workout)
            workout_name = normalize_workout_label(col_a)
            current_workout = {
                "name": workout_name,
                "type": detect_workout_type(workout_name),
                "day_number": len(current_week_workouts) + 1,
                "exercises": [],
            }
            exercise_order = 0

            if col_b:
                exercise_order += 1
                exercise = build_pb3_exercise(row, exercise_order)
                if exercise:
                    current_workout["exercises"].append(exercise)
            continue

        # Regular exercise row
        if col_b and current_workout is not None:
            exercise_order += 1
            exercise = build_pb3_exercise(row, exercise_order)
            if exercise:
                current_workout["exercises"].append(exercise)

    # Flush
    if current_workout is not None:
        current_week_workouts.append(current_workout)
    if current_week_number is not None:
        weeks.append({
            "week_number": current_week_number,
            "workouts": current_week_workouts,
        })

    return {
        "phase_number": 1,
        "name": sheet_name,
        "description": None,
        "weeks": weeks,
    }


def build_pb3_exercise(row, order: int) -> dict | None:
    """Build exercise from PB 3.0 row."""
    def get(idx):
        return clean_value(row[idx]) if len(row) > idx else None

    name = get(1)
    if not name:
        return None

    warmup = get(2)
    working = get(3)
    reps = get(4)
    # col 5 = load (skip), col 6 = %1RM (skip)
    rpe = get(7)
    rest = get(8)
    notes = get(9)

    if working is not None:
        try:
            working = int(working)
        except (ValueError, TypeError):
            pass

    return {
        "order": order,
        "name": str(name),
        "warmup_sets": str(warmup) if warmup is not None else None,
        "working_sets": working,
        "reps": str(reps) if reps is not None else None,
        "rpe": str(rpe) if rpe is not None else None,
        "rest": str(rest) if rest is not None else None,
        "notes": str(notes) if notes is not None else None,
        "substitutions": [],
    }


# ---------------------------------------------------------------------------
# Program definitions
# ---------------------------------------------------------------------------

def extract_ppl_program(filepath: Path, program_name: str, slug: str, frequency: int, description: str) -> dict:
    """Extract a PPL-style program from an Excel file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    phases = []
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        phase = parse_ppl_sheet(ws, sheet_name)
        phase["phase_number"] = i + 1
        phases.append(phase)

    return {
        "name": program_name,
        "slug": slug,
        "frequency": frequency,
        "description": description,
        "source_file": filepath.name,
        "phases": phases,
    }


def extract_pb2_program(filepath: Path, program_name: str, slug: str, frequency: int, description: str) -> dict:
    """Extract a Powerbuilding 2.0 program."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    phases = []
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        phase = parse_pb2_sheet(ws, sheet_name)
        phase["phase_number"] = i + 1
        phases.append(phase)

    return {
        "name": program_name,
        "slug": slug,
        "frequency": frequency,
        "description": description,
        "source_file": filepath.name,
        "phases": phases,
    }


def extract_pb3_program(filepath: Path, program_name: str, slug: str, frequency: int, description: str) -> dict:
    """Extract a Powerbuilding 3.0 program."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    phases = []
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        phase = parse_pb3_sheet(ws, sheet_name)
        phase["phase_number"] = i + 1
        phases.append(phase)

    return {
        "name": program_name,
        "slug": slug,
        "frequency": frequency,
        "description": description,
        "source_file": filepath.name,
        "phases": phases,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

PROGRAMS = [
    {
        "type": "ppl",
        "filepath": JEFF_NIPPARD_DIR / "Ultimate PPL 4x 5x 6x" / "The_Ultimate_Push_Pull_Legs_System_-_4x.xlsx",
        "name": "Ultimate PPL 4x",
        "slug": "ultimate-ppl-4x",
        "frequency": 4,
        "description": "Jeff Nippard's Ultimate Push Pull Legs System - 4x/week",
        "output": "ultimate-ppl-4x.json",
    },
    {
        "type": "ppl",
        "filepath": JEFF_NIPPARD_DIR / "Ultimate PPL 4x 5x 6x" / "The_Ultimate_Push_Pull_Legs_System_-_5x.xlsx",
        "name": "Ultimate PPL 5x",
        "slug": "ultimate-ppl-5x",
        "frequency": 5,
        "description": "Jeff Nippard's Ultimate Push Pull Legs System - 5x/week",
        "output": "ultimate-ppl-5x.json",
    },
    {
        "type": "ppl",
        "filepath": JEFF_NIPPARD_DIR / "Ultimate PPL 4x 5x 6x" / "The_Ultimate_Push_Pull_Legs_System_-_6x.xlsx",
        "name": "Ultimate PPL 6x",
        "slug": "ultimate-ppl-6x",
        "frequency": 6,
        "description": "Jeff Nippard's Ultimate Push Pull Legs System - 6x/week",
        "output": "ultimate-ppl-6x.json",
    },
    {
        "type": "ppl",
        "filepath": JEFF_NIPPARD_DIR / "Ultimate PPL 4x 5x 6x" / "Edited PPL 5x.xlsx",
        "name": "Ultimate PPL 5x (Edited)",
        "slug": "ultimate-ppl-5x-edited",
        "frequency": 5,
        "description": "Jeff Nippard's Ultimate Push Pull Legs System - 5x/week (Edited)",
        "output": "ultimate-ppl-5x-edited.json",
    },
    {
        "type": "pb2",
        "filepath": JEFF_NIPPARD_DIR / "Jeff Nippard - Powerbuilding Version 2" / "POWERBUILDING 2.0 SPREADSHEET 4x.xlsx",
        "name": "Powerbuilding 2.0 4x",
        "slug": "powerbuilding-2-4x",
        "frequency": 4,
        "description": "Jeff Nippard's Powerbuilding 2.0 - 4x/week",
        "output": "powerbuilding-2-4x.json",
    },
    {
        "type": "pb2",
        "filepath": JEFF_NIPPARD_DIR / "Jeff Nippard - Powerbuilding Version 2" / "POWERBUILDING 2.0 SPREADSHEET 5-6x.xlsx",
        "name": "Powerbuilding 2.0 5-6x",
        "slug": "powerbuilding-2-5-6x",
        "frequency": 5,
        "description": "Jeff Nippard's Powerbuilding 2.0 - 5-6x/week",
        "output": "powerbuilding-2-5-6x.json",
    },
    {
        "type": "pb3",
        "filepath": JEFF_NIPPARD_DIR / "Powerbuilding 3.0" / "PowerBuilding 3.0.xlsx",
        "name": "Powerbuilding 3.0 5x",
        "slug": "powerbuilding-3-5x",
        "frequency": 5,
        "description": "Jeff Nippard's Powerbuilding System 3.0 - 5x/week",
        "output": "powerbuilding-3-5x.json",
    },
]


def count_stats(program: dict) -> dict:
    """Count phases, weeks, workouts, and exercises in a program."""
    total_weeks = 0
    total_workouts = 0
    total_exercises = 0
    for phase in program["phases"]:
        for week in phase["weeks"]:
            total_weeks += 1
            for workout in week["workouts"]:
                total_workouts += 1
                total_exercises += len(workout["exercises"])
    return {
        "phases": len(program["phases"]),
        "weeks": total_weeks,
        "workouts": total_workouts,
        "exercises": total_exercises,
    }


def main():
    print(f"Extracting programs to {OUTPUT_DIR}\n")
    print(f"{'Program':<35} {'Phases':>6} {'Weeks':>6} {'Workouts':>9} {'Exercises':>10}  Status")
    print("-" * 80)

    for prog in PROGRAMS:
        filepath = prog["filepath"]
        if not filepath.exists():
            print(f"{prog['name']:<35}  ERROR: file not found: {filepath}")
            continue

        try:
            if prog["type"] == "ppl":
                data = extract_ppl_program(
                    filepath, prog["name"], prog["slug"], prog["frequency"], prog["description"]
                )
            elif prog["type"] == "pb2":
                data = extract_pb2_program(
                    filepath, prog["name"], prog["slug"], prog["frequency"], prog["description"]
                )
            elif prog["type"] == "pb3":
                data = extract_pb3_program(
                    filepath, prog["name"], prog["slug"], prog["frequency"], prog["description"]
                )
            else:
                print(f"{prog['name']:<35}  ERROR: unknown type {prog['type']}")
                continue

            stats = count_stats(data)
            out_path = OUTPUT_DIR / prog["output"]
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)

            concerns = []
            if stats["phases"] == 0:
                concerns.append("NO PHASES")
            if stats["workouts"] == 0:
                concerns.append("NO WORKOUTS")
            if stats["exercises"] == 0:
                concerns.append("NO EXERCISES")

            status = "OK" if not concerns else "WARN: " + ", ".join(concerns)
            print(
                f"{prog['name']:<35} {stats['phases']:>6} {stats['weeks']:>6} {stats['workouts']:>9} {stats['exercises']:>10}  {status}"
            )

        except Exception as e:
            import traceback
            print(f"{prog['name']:<35}  ERROR: {e}")
            traceback.print_exc()

    print("\nDone.")


if __name__ == "__main__":
    main()
